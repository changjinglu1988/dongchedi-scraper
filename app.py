"""Flask 入口。

POST /api/scrape  接收 textarea 多行 URL → 抓取 → 返回多 sheet Excel。
"""

from __future__ import annotations

import io
import re

from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font

from scrapers import get_scraper

app = Flask(__name__)


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "", name).strip() or "未命名车型"
    name = name[:31]
    base, n = name, 2
    while name in used:
        suffix = f"({n})"
        name = (base[: 31 - len(suffix)]) + suffix
        n += 1
    used.add(name)
    return name


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/scrape", methods=["POST"])
def scrape():
    text = (request.form.get("urls") or "").strip()
    if not text:
        return "请输入至少一个车型参数页链接", 400

    urls = [u.strip() for u in text.splitlines() if u.strip()]
    if not urls:
        return "未识别到有效链接", 400

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()
    failed: list[tuple[str, str]] = []
    success_count = 0

    for url in urls:
        try:
            scraper = get_scraper(url)
            tokens = scraper.parse_ids(url)
        except Exception as e:
            failed.append((url, str(e)))
            continue

        for token in tokens:
            try:
                results = scraper.fetch(token)
            except Exception as e:
                failed.append((f"{url} [{token}]", str(e)))
                continue

            for result in results:
                sheet_name = _sanitize_sheet_name(result["model_name"], used_names)
                ws = wb.create_sheet(title=sheet_name)
                ws.append(["类别", "参数名", "参数值"])
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                for r in result["params"]:
                    ws.append([r["category"], r["name"], r["value"]])
                ws.column_dimensions["A"].width = 18
                ws.column_dimensions["B"].width = 32
                ws.column_dimensions["C"].width = 30
                success_count += 1

    if failed:
        ws = wb.create_sheet(title="_失败列表")
        ws.append(["链接", "失败原因"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for url, reason in failed:
            ws.append([url, reason])
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 60

    if success_count == 0:
        detail = "\n".join(f"{u}: {r}" for u, r in failed)
        return f"所有链接抓取失败:\n{detail}", 500

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"车辆参数_{success_count}款.xlsx",
    )


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
